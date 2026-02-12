from pathlib import Path
from typing import Iterator
import logging

from .base import Extractor, ExtractedPage
from .hf_compat import patch_hf_hub

patch_hf_hub()

logger = logging.getLogger(__name__)


class XlsxExtractor(Extractor):
    """
    Extracteur pour les fichiers Excel (xlsx, xls) utilisant Docling.

    Docling préserve la structure sémantique des tableaux en :
    - Détectant automatiquement les en-têtes et hiérarchies
    - Préservant les relations entre cellules
    - Générant une représentation markdown structurée
    - Extrayant les métadonnées et contexte du tableau

    Chaque feuille de calcul est traitée comme une page distincte.
    """
    SUPPORTED_EXTENSIONS = [".xlsx", ".xls"]

    def __init__(self):
        self._converter = None

    def _get_converter(self):
        """Lazy loading du DocumentConverter pour éviter l'import au démarrage."""
        if self._converter is None:
            try:
                from docling.document_converter import DocumentConverter
                self._converter = DocumentConverter()
            except ImportError as e:
                logger.error(
                    "Docling n'est pas installé. Installez-le avec: uv add docling"
                )
                raise ImportError(
                    "docling est requis pour l'extraction Excel. "
                    "Installez-le avec: uv add docling"
                ) from e
        return self._converter

    def extract(self, file_path: Path) -> Iterator[ExtractedPage]:
        """
        Extrait le contenu du fichier Excel en préservant la structure des tableaux.

        Utilise Docling pour :
        - Analyser la structure des feuilles
        - Détecter les tableaux et leurs hiérarchies
        - Convertir en markdown structuré
        """
        try:
            converter = self._get_converter()

            # Conversion du document avec Docling
            result = converter.convert(str(file_path))

            # Docling organise le contenu par pages/sections
            # Pour Excel, chaque feuille est considérée comme une section

            # Extraction du texte complet avec structure préservée
            full_markdown = result.document.export_to_markdown()

            if not full_markdown or not full_markdown.strip():
                logger.warning(f"Aucun contenu extrait de {file_path}")
                return

            # Docling peut séparer par tables, on va essayer d'extraire les tables individuellement
            tables = []
            if hasattr(result.document, 'tables'):
                tables = result.document.tables

            # Si on a des tables explicites, les traiter séparément
            if tables:
                for idx, table in enumerate(tables, start=1):
                    try:
                        # Conversion de la table en markdown
                        table_markdown = table.export_to_markdown() if hasattr(table, 'export_to_markdown') else str(table)

                        # Récupération des métadonnées de la table
                        metadata = {
                            "source_type": "xlsx",
                            "table_index": idx,
                        }

                        # Ajout du nom de la table si disponible
                        if hasattr(table, 'caption') and table.caption:
                            metadata["table_name"] = table.caption

                        # Ajout des dimensions si disponibles
                        if hasattr(table, 'num_rows'):
                            metadata["num_rows"] = table.num_rows
                        if hasattr(table, 'num_cols'):
                            metadata["num_cols"] = table.num_cols

                        yield ExtractedPage(
                            page_number=idx,
                            text=table_markdown,
                            metadata=metadata,
                        )
                    except Exception as e:
                        logger.error(f"Erreur lors de l'extraction de la table {idx} de {file_path}: {e}")
                        continue
            else:
                # Fallback : si pas de tables explicites, utiliser le markdown complet
                # On suppose que c'est une seule "page"
                yield ExtractedPage(
                    page_number=1,
                    text=full_markdown,
                    metadata={
                        "source_type": "xlsx",
                        "extraction_method": "docling_full_document",
                    },
                )

        except Exception as e:
            logger.error(f"Impossible d'extraire le fichier Excel {file_path} avec Docling: {e}")
            # Fallback vers l'ancienne méthode si Docling échoue
            logger.info(f"Tentative de fallback vers openpyxl pour {file_path}")
            try:
                yield from self._fallback_openpyxl_extract(file_path)
            except Exception as fallback_error:
                logger.error(f"Le fallback openpyxl a également échoué: {fallback_error}")
                return

    def _fallback_openpyxl_extract(self, file_path: Path) -> Iterator[ExtractedPage]:
        """
        Méthode de fallback utilisant openpyxl si Docling échoue.
        Conserve l'ancienne logique d'extraction.
        """
        from openpyxl import load_workbook

        wb = None
        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)

            for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
                try:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))

                    if not rows:
                        continue

                    # Utilise la première ligne comme en-têtes
                    headers = [
                        str(header) if header is not None else f"col_{index}"
                        for index, header in enumerate(rows[0])
                    ]

                    # Construit un préambule contextuel
                    context_header = (
                        f"=== Feuille Excel: {sheet_name} ===\n"
                        f"Nombre de lignes: {len(rows) - 1}\n"
                        f"Colonnes: {', '.join(headers)}\n"
                        f"--- Données ---"
                    )

                    lines = []
                    for row in rows[1:]:
                        parts = []
                        for header, val in zip(headers, row):
                            if val is not None and str(val).strip():
                                parts.append(f"{header}: {val}")
                        if parts:
                            lines.append("; ".join(parts))

                    if lines:
                        text = context_header + "\n\n" + "\n".join(lines)
                        yield ExtractedPage(
                            page_number=sheet_idx,
                            text=text,
                            metadata={
                                "source_type": "xlsx",
                                "sheet_name": sheet_name,
                                "extraction_method": "openpyxl_fallback",
                            },
                        )

                except Exception as e:
                    logger.error(f"Erreur lors de l'extraction de la feuille '{sheet_name}': {e}")
                    continue
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception as e:
                    logger.warning(f"Erreur lors de la fermeture du workbook: {e}")
